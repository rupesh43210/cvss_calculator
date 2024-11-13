import openpyxl
import logging
from datetime import datetime
import os
from pathlib import Path

class ExcelProcessor:
    def __init__(self, input_file, gui=None):
        self.input_file = input_file
        self.gui = gui
        self.wb = None
        self.sheet = None
        self.output_file = None

    def log_progress(self, message, progress=None):
        """Log progress to both file and GUI if available"""
        logging.info(message)
        if self.gui:
            self.gui.update_status(message)
            if progress is not None:
                self.gui.update_progress(progress)

    def process_file(self, vector_generator, cvss_calculator):
        """Main method to process Excel file"""
        try:
            # Load workbook
            self.log_progress("Loading workbook...", 10)
            self.wb = openpyxl.load_workbook(self.input_file)
            self.sheet = self.wb.active

            # Get headers
            headers = [str(cell.value).strip() if cell.value else "" for cell in self.sheet[1]]
            self.log_progress(f"Found columns: {headers}", 20)

            # Identify threat description column
            threat_col = self.identify_column(headers)
            if not threat_col:
                raise ValueError(f"Could not identify threat column. Available columns: {headers}")

            # Create output file
            self.create_output_file()

            # Add result columns
            start_col = len(headers) + 1
            self.add_result_columns(start_col)

            # Process each row
            total_rows = self.sheet.max_row - 1  # Excluding header
            successful = 0
            failed = 0

            for row in range(2, self.sheet.max_row + 1):
                # Calculate progress percentage
                progress = 20 + (row - 1) / (self.sheet.max_row - 1) * 70
                
                threat_desc = self.sheet.cell(row=row, column=threat_col).value
                if not threat_desc:
                    continue

                self.log_progress(f"Processing row {row-1}/{total_rows}", progress)
                self.log_progress(f"Analyzing threat: {threat_desc[:100]}...")

                try:
                    # Generate CVSS vector
                    vector = vector_generator.generate_vector(threat_desc)
                    if vector:
                        # Calculate CVSS score
                        result = cvss_calculator.calculate_score(vector)
                        self.save_results(row, start_col, result)
                        successful += 1
                    else:
                        self.sheet.cell(row=row, column=start_col + 4, value="Could not generate vector")
                        failed += 1
                except Exception as e:
                    self.log_progress(f"Error processing row {row}: {str(e)}")
                    failed += 1

            # Save workbook
            self.log_progress("Saving results...", 90)
            self.wb.save(self.output_file)

            # Final summary
            summary = f"""
            Processing complete:
            Total threats processed: {total_rows}
            Successful calculations: {successful}
            Failed calculations: {failed}
            Results saved to: {self.output_file}
            """
            self.log_progress(summary, 100)

            return self.output_file

        except Exception as e:
            error_msg = f"Error processing file: {str(e)}"
            self.log_progress(error_msg)
            raise

    def identify_column(self, headers):
        """Identify threat description column"""
        possible_names = [
            'threat description', 'threat', 'description', 'threats', 
            'vulnerability', 'risk description', 'risk', 'finding', 
            'threat scenario', 'scenario description'
        ]

        for idx, header in enumerate(headers, 1):
            header_lower = header.lower()
            if any(name in header_lower for name in possible_names):
                self.log_progress(f"Using column: '{header}' for threat descriptions")
                return idx
        return None

    def create_output_file(self):
        """Create output filename with timestamp"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = os.path.dirname(self.input_file)
        base_name = os.path.basename(self.input_file)
        self.output_file = os.path.join(output_dir, f"cvss_scored_{timestamp}_{base_name}")

    def add_result_columns(self, start_col):
        """Add result columns to worksheet"""
        headers = [
            'CVSS Vector', 'Base Score', 'Severity', 'Temporal Score',
            'Environmental Score', 'Attack Vector', 'Attack Complexity',
            'Privileges Required', 'User Interaction', 'Impact Scores'
        ]
        
        for idx, header in enumerate(headers):
            self.sheet.cell(row=1, column=start_col + idx, value=header)

    def save_results(self, row, start_col, results):
        """Save calculation results to Excel"""
        try:
            # Save vector string and scores
            self.sheet.cell(row=row, column=start_col, value=results['vector_string'])
            self.sheet.cell(row=row, column=start_col + 1, value=results['base_score'])
            self.sheet.cell(row=row, column=start_col + 2, value=results['base_severity'])
            
            # Save temporal score if available
            if 'temporal_score' in results and results['temporal_score']:
                self.sheet.cell(row=row, column=start_col + 3, value=results['temporal_score'])
            
            # Save environmental score if available
            if 'environmental_score' in results and results['environmental_score']:
                self.sheet.cell(row=row, column=start_col + 4, value=results['environmental_score'])
            
            # Save individual metrics if available
            metrics = self.parse_vector_metrics(results['vector_string'])
            col_offset = 5
            for value in metrics.values():
                self.sheet.cell(row=row, column=start_col + col_offset, value=value)
                col_offset += 1

        except Exception as e:
            self.log_progress(f"Error saving results for row {row}: {str(e)}")

    @staticmethod
    def parse_vector_metrics(vector_string):
        """Parse individual metrics from vector string"""
        metrics = {}
        parts = vector_string.split('/')
        for part in parts:
            if ':' in part:
                key, value = part.split(':')
                metrics[key] = value
        return metrics