import openpyxl
import requests
from time import sleep
import re
from datetime import datetime
import logging
import sys
from pathlib import Path
import spacy
from collections import defaultdict
import tkinter as tk
from tkinter import filedialog, messagebox
import os

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('cvss_calculator.log'),
        logging.StreamHandler(sys.stdout)
    ]
)

class CVSSCalculatorGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("CVSS Calculator")
        self.root.geometry("600x400")
        self.setup_gui()

    def setup_gui(self):
        # File selection
        frame = tk.Frame(self.root, padx=20, pady=20)
        frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(frame, text="CVSS Score Calculator", font=('Arial', 14, 'bold')).pack(pady=10)
        
        tk.Button(frame, text="Select Excel File", command=self.select_file).pack(pady=20)
        
        # Status display
        self.status_var = tk.StringVar()
        self.status_var.set("Please select an Excel file to begin...")
        tk.Label(frame, textvariable=self.status_var, wraplength=500).pack(pady=10)

        # Progress display
        self.progress_var = tk.StringVar()
        tk.Label(frame, textvariable=self.progress_var).pack(pady=10)

    def select_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")],
            title="Select Excel file containing threat descriptions"
        )
        if file_path:
            self.status_var.set(f"Processing file: {file_path}")
            self.root.update()
            try:
                calculator = AutomatedCVSSCalculator(file_path, self)
                calculator.process_excel()
                messagebox.showinfo("Success", "Processing complete! Check the output file and logs for details.")
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred: {str(e)}")
            self.status_var.set("Ready for next file...")

    def update_progress(self, message):
        self.progress_var.set(message)
        self.root.update()

    def run(self):
        self.root.mainloop()

class AutomatedCVSSCalculator:
    def __init__(self, input_file, gui=None):
        self.input_file = input_file
        self.gui = gui
        self.wb = None
        self.sheet = None
        self.output_file = None
        self.api_base_url = "https://cvss.nist.gov/api/score/vector/"
        
        # Load NLP model
        self.log_progress("Loading NLP model...")
        try:
            self.nlp = spacy.load("en_core_web_md")
        except OSError:
            self.log_progress("Downloading spaCy model...")
            spacy.cli.download("en_core_web_md")
            self.nlp = spacy.load("en_core_web_md")

        # CVSS metric determination rules
        self.metric_rules = {
            'AV': {  # Attack Vector
                'keywords': {
                    'N': ['internet', 'remote', 'network', 'web', 'online', 'external', 'internet-facing'],
                    'A': ['adjacent', 'local network', 'lan', 'neighbor', 'neighbouring'],
                    'L': ['local', 'physical access', 'locally', 'system access'],
                    'P': ['physical', 'hardware', 'device', 'physically']
                },
                'default': 'N'
            },
            'AC': {  # Attack Complexity
                'keywords': {
                    'L': ['simple', 'easily', 'straightforward', 'common', 'known vulnerability'],
                    'H': ['complex', 'difficult', 'sophisticated', 'chain', 'multiple steps', 'specific condition']
                },
                'default': 'L'
            },
            'PR': {  # Privileges Required
                'keywords': {
                    'N': ['unauthenticated', 'no authentication', 'anonymous', 'without login'],
                    'L': ['authenticated', 'basic user', 'normal user', 'user account'],
                    'H': ['administrative', 'admin', 'privileged', 'root', 'system level']
                },
                'default': 'N'
            },
            'UI': {  # User Interaction
                'keywords': {
                    'N': ['automatic', 'without user', 'no interaction', 'automated'],
                    'R': ['user action', 'click', 'download', 'user interaction', 'manual']
                },
                'default': 'N'
            },
            'S': {  # Scope
                'keywords': {
                    'U': ['single system', 'same system', 'unchanged', 'contained'],
                    'C': ['multiple systems', 'spread', 'other systems', 'changed', 'escalate']
                },
                'default': 'U'
            },
            'C': {  # Confidentiality
                'keywords': {
                    'H': ['sensitive data', 'credentials', 'passwords', 'full access', 'all data'],
                    'L': ['limited information', 'partial disclosure', 'minor'],
                    'N': ['no confidentiality', 'no data disclosure']
                },
                'default': 'L'
            },
            'I': {  # Integrity
                'keywords': {
                    'H': ['modify all', 'complete corruption', 'full control'],
                    'L': ['minor modification', 'partial modification', 'slight changes'],
                    'N': ['no integrity', 'read only', 'no modification']
                },
                'default': 'L'
            },
            'A': {  # Availability
                'keywords': {
                    'H': ['crash', 'denial of service', 'dos', 'shutdown', 'unavailable'],
                    'L': ['degraded', 'intermittent', 'reduced performance'],
                    'N': ['no availability', 'no impact on availability']
                },
                'default': 'L'
            }
        }

    def log_progress(self, message):
        logging.info(message)
        if self.gui:
            self.gui.update_progress(message)

    def analyze_threat_description(self, threat_desc):
        """Analyze threat description using NLP to determine appropriate CVSS metrics"""
        if not threat_desc or not isinstance(threat_desc, str):
            return None

        # Process the text with spaCy
        doc = self.nlp(threat_desc.lower())
        
        # Initialize scores for each metric option
        metric_scores = defaultdict(lambda: defaultdict(float))
        
        # Analyze text and score each metric option
        for metric, rules in self.metric_rules.items():
            for value, keywords in rules['keywords'].items():
                score = 0
                for keyword in keywords:
                    # Check for exact matches
                    if keyword in doc.text:
                        score += 1
                    
                    # Check for semantic similarity with key phrases
                    keyword_doc = self.nlp(keyword)
                    for chunk in doc.noun_chunks:
                        similarity = keyword_doc.similarity(chunk)
                        if similarity > 0.7:
                            score += similarity
                
                metric_scores[metric][value] = score
        
        # Determine final metrics based on highest scores or defaults
        final_metrics = {}
        for metric, scores in metric_scores.items():
            if scores:
                max_score_value = max(scores.items(), key=lambda x: x[1])[0]
                final_metrics[metric] = max_score_value
            else:
                final_metrics[metric] = self.metric_rules[metric]['default']
        
        # Construct CVSS vector string
        vector = f"CVSS:3.1/AV:{final_metrics['AV']}/AC:{final_metrics['AC']}/PR:{final_metrics['PR']}/"
        vector += f"UI:{final_metrics['UI']}/S:{final_metrics['S']}/C:{final_metrics['C']}/"
        vector += f"I:{final_metrics['I']}/A:{final_metrics['A']}"
        
        return vector

    def calculate_cvss_score(self, vector):
        """Calculate CVSS score using NIST calculator API"""
        try:
            encoded_vector = requests.utils.quote(vector)
            response = requests.get(f"{self.api_base_url}{encoded_vector}", timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                return {
                    'score': data['baseScore'],
                    'severity': data['baseSeverity'],
                    'vector': vector
                }
            else:
                return {'error': f'API Error: {response.status_code}'}
                
        except Exception as e:
            return {'error': f'Error calculating score: {str(e)}'}

    def process_excel(self):
        """Process the Excel file and calculate CVSS scores"""
        try:
            self.log_progress(f"Loading workbook: {self.input_file}")
            self.wb = openpyxl.load_workbook(self.input_file)
            self.sheet = self.wb.active

            # Create output filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_dir = os.path.dirname(self.input_file)
            base_name = os.path.basename(self.input_file)
            self.output_file = os.path.join(output_dir, f"cvss_scored_{timestamp}_{base_name}")

            # Find header row and column indices
            headers = [cell.value for cell in self.sheet[1]]
            threat_desc_col = next((i + 1 for i, h in enumerate(headers) 
                                  if h and 'threat' in h.lower() and 'description' in h.lower()), None)
            
            if not threat_desc_col:
                raise ValueError("Required column 'Threat Description' not found")

            # Add new columns for results
            vector_col = len(headers) + 1
            score_col = vector_col + 1
            severity_col = score_col + 1
            error_col = severity_col + 1

            self.sheet.cell(row=1, column=vector_col, value='CVSS Vector')
            self.sheet.cell(row=1, column=score_col, value='CVSS Score')
            self.sheet.cell(row=1, column=severity_col, value='Severity')
            self.sheet.cell(row=1, column=error_col, value='Error')

            # Process each row
            total_rows = self.sheet.max_row - 1
            successful = 0
            failed = 0

            for row in range(2, self.sheet.max_row + 1):
                threat_desc = self.sheet.cell(row=row, column=threat_desc_col).value

                if not threat_desc:
                    continue

                self.log_progress(f"Processing row {row-1}/{total_rows}")

                # Generate CVSS vector from threat description
                vector = self.analyze_threat_description(threat_desc)
                if vector:
                    self.sheet.cell(row=row, column=vector_col, value=vector)
                    
                    # Calculate CVSS score
                    result = self.calculate_cvss_score(vector)
                    
                    if 'error' in result:
                        self.sheet.cell(row=row, column=error_col, value=result['error'])
                        failed += 1
                    else:
                        self.sheet.cell(row=row, column=score_col, value=result['score'])
                        self.sheet.cell(row=row, column=severity_col, value=result['severity'])
                        successful += 1
                else:
                    self.sheet.cell(row=row, column=error_col, value="Could not generate CVSS vector")
                    failed += 1

                sleep(0.5)  # API rate limiting

            # Save results
            self.wb.save(self.output_file)
            
            summary = f"""
            Processing complete:
            Total threats processed: {total_rows}
            Successful calculations: {successful}
            Failed calculations: {failed}
            Results saved to: {self.output_file}
            """
            self.log_progress(summary)

        except Exception as e:
            error_msg = f"Error processing file: {str(e)}"
            logging.error(error_msg)
            raise Exception(error_msg)

def main():
    # Create and run GUI
    gui = CVSSCalculatorGUI()
    gui.run()

if __name__ == "__main__":
    main()
